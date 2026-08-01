import json
import re
from copy import deepcopy
from pathlib import Path

import openpyxl


WORKBOOK_PATH = Path(r"C:\Users\gradp\Downloads\AP_WESELE.xlsx")
SNAPSHOT_PATH = Path("supabase/ap-wesele-snapshot.json")
SEED_PATH = Path("supabase/seed-production.sql")
WEDDING_ID = "aaaaaaaa-2028-4000-8000-000000000001"
WEDDING_SLUG = "aleksandra-pawel-2028"


BASE_SNAPSHOT = {
    "wedding": {
        "bride": "Aleksandra",
        "groom": "Pawel",
        "date": "2028-06-17",
        "ceremonyTime": "14:00",
        "ceremonyAddress": "Miejsce ceremonii do uzupelnienia",
        "venueAddress": "Sala weselna do uzupelnienia",
        "welcomeText": "Cieszymy sie, ze bedziesz z nami. Tu znajdziesz najwazniejsze informacje o naszym weselu.",
        "contactPhone": "+48 500 100 200",
        "transportInfo": "Szczegoly transportu uzupelnimy blizej wesela.",
    },
    "schedule": [
        {"id": "schedule-ceremony", "time": "14:00", "title": "Ceremonia slubna", "place": "Do uzupelnienia", "owner": "Aleksandra i Pawel", "status": "planned"},
        {"id": "schedule-party", "time": "16:00", "title": "Przyjecie weselne", "place": "Do uzupelnienia", "owner": "Aleksandra i Pawel", "status": "planned"},
        {"id": "schedule-cake", "time": "20:00", "title": "Tort weselny", "place": "Sala weselna", "owner": "Do uzupelnienia", "status": "planned"},
    ],
    "guests": [],
    "tables": [],
    "roomElements": [
        {"id": "dance", "label": "Parkiet", "type": "dance", "x": 40, "y": 26, "w": 20, "h": 16},
        {"id": "entry", "label": "Wejscie", "type": "entry", "x": 6, "y": 78, "w": 14, "h": 10},
    ],
    "faqItems": [
        {"id": "faq-rsvp", "question": "Do kiedy potwierdzic obecnosc?", "answer": "Prosimy o odpowiedz do 31 maja 2028.", "active": True},
        {"id": "faq-photos", "question": "Gdzie dodawac zdjecia?", "answer": "Uzyj przycisku Dodaj zdjecia albo kodu QR.", "active": True},
    ],
    "qrInvites": [
        {"id": "qr-rsvp", "label": "RSVP", "target": "/rsvp", "scans": 0, "active": True},
        {"id": "qr-gallery", "label": "Dodaj zdjecia", "target": "/upload", "scans": 0, "active": True},
    ],
    "gallery": [],
    "planning": {
        "budgetTarget": 0,
        "tasks": [],
        "vendors": [],
        "expenses": [],
        "payments": [],
        "documents": [],
        "attachments": [],
    },
    "theme": {
        "coupleName": "Aleksandra i Pawel",
        "themeId": "gold",
        "accentColor": "#2f7d6d",
        "coverStyle": "editorial",
        "accessMode": "code",
        "weddingCode": "AP2028",
        "publicRsvp": True,
        "galleryModeration": True,
        "showWholeRoomToGuests": False,
    },
}


def text(value):
    if value is None:
        return ""
    return str(value).strip()


def number(value):
    if value is None or value == "":
        return 0
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0


def slugify(value):
    value = text(value).lower()
    value = re.sub(r"[^a-z0-9ąćęłńóśźż]+", "-", value, flags=re.IGNORECASE)
    value = value.strip("-").lower()
    return value or "item"


def split_name(full_name):
    parts = text(full_name).split()
    if not parts:
        return "", ""
    if len(parts) == 1:
        return parts[0], ""
    return parts[0], " ".join(parts[1:])


def group_label(raw):
    value = text(raw).upper()
    if value == "O":
        return "Strona Aleksandry (O)"
    if value == "P":
        return "Strona Pawla (P)"
    if value == "O/P":
        return "Wspolni (O/P)"
    return text(raw) or "Do ustalenia"


def expense_category(raw):
    value = text(raw).lower()
    if "sala" in value:
        return "venue"
    if "muzyk" in value:
        return "music"
    if "foto" in value or "wideo" in value:
        return "photo"
    if "dekor" in value or "kwiat" in value:
        return "decor"
    if "alkohol" in value or "napoje" in value or "słodkie" in value or "slodkie" in value:
        return "food"
    if "transport" in value or "auto" in value:
        return "transport"
    if "ubran" in value or "obrącz" in value or "obracz" in value:
        return "outfit"
    if "papier" in value or "zaprosz" in value:
        return "paper"
    return "other"


def make_tables(guest_count):
    tables = [{"id": "table-1", "number": 1, "name": "Stol prezydialny", "shape": "head", "capacity": 10, "x": 50, "y": 14, "theme": ""}]
    needed_regular = max(0, guest_count - 10)
    regular_tables = max(1, (needed_regular + 7) // 8)
    positions = [
        (18, 34), (38, 34), (62, 34), (82, 34),
        (18, 54), (38, 54), (62, 54), (82, 54),
        (18, 74), (38, 74), (62, 74), (82, 74),
        (28, 88), (50, 88), (72, 88),
    ]
    for index in range(regular_tables):
        x, y = positions[index % len(positions)]
        tables.append({
            "id": f"table-{index + 2}",
            "number": index + 2,
            "name": f"Stolik {index + 2}",
            "shape": "round",
            "capacity": 8,
            "x": x,
            "y": y,
            "theme": "",
        })
    return tables


def read_guests(wb):
    ws = wb.worksheets[1]
    guests = []
    for row in ws.iter_rows(min_row=8, max_row=ws.max_row, values_only=True):
        name = text(row[1] if len(row) > 1 else "")
        if not name:
            continue
        index = len(guests) + 1
        first_name, last_name = split_name(name)
        rate = number(row[3] if len(row) > 3 else None)
        menu_cost = number(row[4] if len(row) > 4 else None)
        note = text(row[5] if len(row) > 5 else "")
        child = rate > 0 and rate < 1 or "dziecko" in name.lower()
        dietary = []
        if rate > 0 and rate < 1:
            dietary.append("Stawka menu 70%")
        if note:
            dietary.append(note)
        guests.append({
            "id": f"guest-{index:03d}",
            "firstName": first_name,
            "lastName": last_name,
            "companion": "",
            "group": group_label(row[2] if len(row) > 2 else ""),
            "status": "invited",
            "tableId": "",
            "seat": 1,
            "dietaryNotes": "; ".join(dietary),
            "child": child,
            "accommodation": "",
            "transport": False,
            "note": f"Excel: grupa {text(row[2] if len(row) > 2 else '')}, stawka {rate:g}, koszt menu {menu_cost:g} zl".strip(),
            "token": f"ap2028-guest-{index:03d}",
        })
    return guests


def read_budget(wb):
    ws = wb.worksheets[0]
    vendors = []
    expenses = []
    payments = []
    vendor_by_name = {}

    budget_target = int(number(ws.cell(row=3, column=1).value))

    for row in ws.iter_rows(min_row=5, max_row=47, values_only=True):
        category_raw = text(row[1] if len(row) > 1 else "")
        label = text(row[2] if len(row) > 2 else "")
        if not label or label.lower() == "pozycja / usługa":
            continue
        vendor_name = text(row[3] if len(row) > 3 else "") or label
        notes = text(row[4] if len(row) > 4 else "")
        qty = number(row[5] if len(row) > 5 else None)
        unit = text(row[6] if len(row) > 6 else "")
        unit_price = number(row[7] if len(row) > 7 else None)
        planned = number(row[8] if len(row) > 8 else None)
        actual = number(row[9] if len(row) > 9 else None)
        deposit = number(row[10] if len(row) > 10 else None)
        remaining = number(row[11] if len(row) > 11 else None)
        amount = actual or planned
        paid_amount = min(deposit, amount) if amount else deposit

        vendor_key = vendor_name.lower()
        if vendor_key not in vendor_by_name:
            vendor_id = f"vendor-{len(vendors) + 1:03d}"
            vendor_by_name[vendor_key] = vendor_id
            vendors.append({
                "id": vendor_id,
                "category": category_raw or "Inne",
                "name": vendor_name,
                "contactName": "",
                "phone": "",
                "email": "",
                "status": "booked" if deposit > 0 else "shortlisted",
                "contractStatus": "missing",
                "totalCost": amount,
                "depositPaid": paid_amount,
                "paymentDueDate": "2028-06-01",
                "notes": notes,
            })
        vendor_id = vendor_by_name[vendor_key]

        payment_id = ""
        if deposit > 0:
            payment_id = f"payment-{len(payments) + 1:03d}"
            payments.append({
                "id": payment_id,
                "vendorId": vendor_id,
                "label": f"Zaliczka: {label}",
                "amount": deposit,
                "dueDate": "2026-08-01",
                "paid": True,
                "method": "do uzupelnienia",
            })
        if remaining > 0:
            payments.append({
                "id": f"payment-{len(payments) + 1:03d}",
                "vendorId": vendor_id,
                "label": f"Do zaplaty: {label}",
                "amount": remaining,
                "dueDate": "2028-06-01",
                "paid": False,
                "method": "do uzupelnienia",
            })

        status = "paid" if amount > 0 and paid_amount >= amount else "deposit-paid" if paid_amount > 0 else "planned"
        expense_note = notes
        details = []
        if qty:
            details.append(f"Ilosc: {qty:g} {unit}".strip())
        if unit_price:
            details.append(f"Cena jedn.: {unit_price:g} zl")
        if details:
            expense_note = f"{expense_note}; {'; '.join(details)}".strip("; ")
        expenses.append({
            "id": f"expense-{len(expenses) + 1:03d}",
            "label": label,
            "category": expense_category(category_raw),
            "vendorId": vendor_id,
            "paymentId": payment_id,
            "documentId": "",
            "amount": amount,
            "paidAmount": paid_amount,
            "dueDate": "2028-06-01",
            "status": status,
            "fileName": "",
            "imageName": "",
            "note": expense_note,
        })

    return budget_target, vendors, expenses, payments


def sql_literal_json(data):
    return json.dumps(data, ensure_ascii=False, indent=2).replace("'", "''")


def build_seed(snapshot):
    data_json = sql_literal_json(snapshot)
    return f"""-- Clean production seed for Aleksandra & Pawel.
-- Generated from AP_WESELE.xlsx.
-- Run after all migrations in Supabase SQL Editor.

delete from public.weddings
where slug in ('{WEDDING_SLUG}', 'nasze-wesele');

insert into public.weddings (
  id,
  slug,
  couple_names,
  wedding_date,
  ceremony_time,
  reception_time,
  ceremony_location,
  reception_location,
  hero_message,
  rsvp_deadline,
  contact_email,
  contact_phone,
  is_published,
  owner_email,
  public_url,
  plan_id,
  storage_limit_mb,
  video_limit_minutes,
  expires_at,
  privacy_level
) values (
  '{WEDDING_ID}',
  '{WEDDING_SLUG}',
  'Aleksandra i Pawel',
  '2028-06-17',
  '14:00',
  '16:00',
  'Miejsce ceremonii do uzupelnienia',
  'Sala weselna do uzupelnienia',
  'Cieszymy sie, ze bedziesz z nami. Tu znajdziesz najwazniejsze informacje o naszym weselu.',
  '2028-05-31',
  'kontakt@aleksandrapawel-2028.pl',
  '+48 500 100 200',
  true,
  'kontakt@aleksandrapawel-2028.pl',
  '/',
  'live',
  10240,
  30,
  '2028-12-31 23:59:59+01',
  'wedding_code'
);

insert into public.wedding_admin_snapshots (wedding_id, data)
values (
  '{WEDDING_ID}',
  '{data_json}'::jsonb
)
on conflict (wedding_id)
do update set data = excluded.data, updated_at = now();
"""


def main():
    wb = openpyxl.load_workbook(WORKBOOK_PATH, data_only=True)
    snapshot = deepcopy(BASE_SNAPSHOT)
    guests = read_guests(wb)
    budget_target, vendors, expenses, payments = read_budget(wb)
    snapshot["guests"] = guests
    snapshot["tables"] = make_tables(len(guests))
    snapshot["planning"] = {
        "budgetTarget": budget_target,
        "tasks": [],
        "vendors": vendors,
        "expenses": expenses,
        "payments": payments,
        "documents": [],
        "attachments": [],
    }

    SNAPSHOT_PATH.write_text(json.dumps(snapshot, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    SEED_PATH.write_text(build_seed(snapshot), encoding="utf-8")
    print(json.dumps({
        "guests": len(guests),
        "tables": len(snapshot["tables"]),
        "budgetTarget": budget_target,
        "vendors": len(vendors),
        "expenses": len(expenses),
        "payments": len(payments),
        "snapshot": str(SNAPSHOT_PATH),
        "seed": str(SEED_PATH),
    }, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
